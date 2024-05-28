# 目标网址： https://bugzilla.yoctoproject.org/buglist.cgi?bug_status=__all__&limit=0&no_redirect=1&order=changeddate%20DESC%2Cpriority%2Cbug_severity&query_format=specific
from bs4 import BeautifulSoup
import lxml
import requests
import re
import pandas as pd
import openpyxl
import os

targetUrl = "https://bugzilla.yoctoproject.org/buglist.cgi?bug_status=__all__&limit=0&no_redirect=1&order=changeddate%20DESC%2Cpriority%2Cbug_severity&query_format=specific"
# targetUrl = "https://bugzilla.yoctoproject.org/buglist.cgi?bug_status=__open__&no_redirect=1&order=changeddate%20DESC%2Cpriority%2Cbug_severity&query_format=specific"

# from selenium import webdriver
# from selenium.webdriver.common.by import By
# browser = webdriver.Chrome(options=webdriver.ChromeOptions())
Headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0',
    'Cookie': 'DEFAULTFORMAT=specific; BUGLIST=15463-14946-15245-15391-15388-15336-15332-8715-15380-15395-6478-13411-13376-14937-15471-15470-14125-15227-10452-15466-15403-15079-15268-15337-15291-15254-10422-13896-13878-7636-12634-7600-7037-15454-15469-15467-15244-15468-15436-15460-15318-15237-15185-15224-15098-15423-15462-4120-9651-15398-15333-15239-4530-12678-15162-10820-15383-15390-14885-15341-15453-15265-15263-15446-15464-15286-15465-14165-14311-15335-15451-15372-14464-8760-10728-15172-14710-11704-14007-15444-14989-15270-15342-15288-15101-15352-14787-15459-15455-15180-14931-15343-15359-15273-14126-15364-15061-14403-13766-13190-15071-15328-14430-14443-15316-14693-14960-15059-14681-14044-13931-13360-13226-8191-13052-14923-5305-9233-14372-14290-15448-15320-15357-15457-15299-14891-15417-15419-15421-15407-13424-15213-14620-13808-15102-14717-14870-14811-11899-15276-15130-11781-13599-15441-12904-15443-15410-15408-12482-2951-14905-14876-15409-15084-14667-14995-13338-10731-14066-15363-15031-8729-13206-11385-15402-4063-5709-15404-15415-15442-15439-15438-14410-14503-13225-15218-13589-14616-5600-6158-15345-14670-15111-14830-9612-13835-14432-14368-13417-15107-14802-12688-14747-14121-14534-14543-14873-8909-14509-14835-15106-14377-14498-14436-14535-14357-13103-14814-13111-4284-13533-13109-14758-13173-13626-15267-12723-13620-13621-14370-13000-15430-12937-12862-4836-15432-15435-15434-15433-15431-15429-15427-15426-15425-15424-15422-15418-15416-14380-15014-14386-15120-14098-14523-13288-14903-14781-14930-13285-15327-12963-10061-12290-15055-13181-14139-14461-14462-14843-13868-13674-14520-14118-14528-13920-14020-14749-12760-13910-14348-14218-14866-8677-15087-15095-14804-14206-14155-14154-15204-15020-5876-14303-14945-5322-13980-13330-10693-6428-14584-14196-13004-12292-15360-14944-15241-14136-3433-9391-4011-15329-9589-9588-9643-15039-15188-15406-15405-15319-14736-15294-13508-15399-15394-12086-12922-14072-15396-15340-15258-14235-6434-12046-15386-15385-15156-15259-10322-15371-14956-6150-1658-6649-5955-14909-14907-15377-15375-15370-12599-12130-11195-15225-10847-5421-15269-15264-15260-15257-15256-13034-12636-15080-12279-14986-7792-15330-13909-13613-13251-13928-15346-14268-14572-15145-14723-13938-15169-15066-15021-14867-15358-15356-15355-15354-15353-15351-9883-9005-9153-8577-15339-15338-9442-13979-10722-13846-13780-15326-5389-15325-15012-14654-9617-13245-5255-7747-7751-8218-8366-7742-7294-7727-7748-6095-6840-7219-7038-6897-9812-9895-8488-8790-8806-8866-8896-8955-8959-9109-9164-8807-9519-9684-9914-9969-8426-8326-8858-8860-8859-8857-9297-9485-9518-9141-9118-9435-9209-9670-9856-9977-10205-8425-8424-10281-9117-8434-8248-8865-8685-14935-14401-13908-9839-15305-15304-15307-14564-15302-15301-15298-15297-15287-15076-12296-6372-14316-13757-15037-11546-11458-11398-11397-14962-13949-15126-15109-13271-14963-14966-11528-14852-14148-14704-14137-13521-13263-10954-10749-10737-14576-14021-13917-13679-13378-13218-13515-13259-15280-6308; LASTORDER=changeddate%20DESC%2Cpriority%2Cbug_severity; Bugzilla_login=3214; Bugzilla_logincookie=DxEgH1r2SU'
}# 这里应该是cookie起效果，需要时常更新cookie

# def login_bugzilla(): # 模拟登录，以便访问所有信息  ——失败(找不到登录的请求链接和请求参数)
#     loginInfo = {
#         'Bugzilla_login': 'zcj15984304723@163.com',
#         'Bugzilla_password': '135790zcjZCJ',
#         'GoAheadAndLogIn': 'Log in',
#         'Bugzilla_login_token': '1713680718-BHXeLxodmtpCT1Oub7rt2RTUQveBf9qpvOGdsFZschs'
#     }
#     url = "https://bugzilla.yoctoproject.org/index.cgi"
#     session = requests.Session()
#     response = session.post(url, params=loginInfo, headers=Headers)
#     if response.status_code == 200:
#         print("登录成功")
#     else:
#         print("登录失败")
#     return session

#     browser.get('https://bugzilla.yoctoproject.org/index.cgi?GoAheadAndLogIn=1') # selenuim模拟登录——失败
#     input = browser.find_element(By.ID,'Bugzilla_login')
#     input.send_keys('')
#     input = browser.find_element(By.ID,'Bugzilla_password')
#     input.send_keys('')
#     submit = browser.find_element(By.ID,'login_in')
#     submit.click() # 点击登录按钮
#     browser.get("https://bugzilla.yoctoproject.org/show_activity.cgi?id=15391")
#     try:
#         print(browser.page_source)  # 输出网页源码
#     except Exception as e:
#         print(str(e))
# login_bugzilla()

# 全局记录条数
reportedId =0
modifiedId =0
commentId =0
modified_data = []
reported_data =[]

def get_html(url):
    try:
        response = requests.get(url, headers=Headers)
        response.raise_for_status()  # 检查请求是否成功
        html = response.text
        soup = BeautifulSoup(html, 'lxml')
        return soup
    except requests.exceptions.RequestException as e:
        # 处理请求异常
        print(f"请求错误: {e}")
    except Exception as e:
        # 处理其他异常
        print(f"发生错误: {e}")


def get_table(soup):
    table = soup.find('table', attrs={'class': 'bz_buglist'})
    return table

def get_trs(table):
    trs = table.find_all('tr', attrs={'class': 'bz_bugitem'})
    return trs

def analyse_tr(tr):
    tds = tr.find_all('td')
    if len(tds) >6:
        bug_id = tds[0].text.strip()
        bug_product = tds[1].text.strip()
        bug_component = tds[2].text.strip()
        bug_assignee = tds[3].text.strip()
        bug_status = tds[4].text.strip()
        bug_summary = tds[6].text.strip()
        return [bug_id, bug_product, bug_component, bug_assignee, bug_status, bug_summary]
    else:
        return []

def get_bugInfo_table(soup):
    table = soup.find('table', attrs={'class': 'edit_form'})
    return table

def get_modified_history(bugId):
    global modifiedId  # 声明 modifiedId 为全局变量
    global modified_data  # 声明 modified_data 为全局变量

    url = f"https://bugzilla.yoctoproject.org/show_activity.cgi?id={bugId}"
    soup = get_html(url)
    if soup:
        table = soup.find('table', attrs={'id': 'bug_activity'})
        if table:
            trs = table.find_all('tr') # type: ignore
            modified = []
            for tr in trs:
                tds = tr.find_all('td')
                if len(tds) == 5:
                    modified_tmp = [modifiedId,tds[0].text.strip(),tds[1].text.strip(),bugId] # 0是人 ，1是时间
                    modified.append(modified_tmp)
                    modifiedId+=1
            return modified
        else:
            print("获取bug修改历史失败")
    else:
        print("获取页面失败")

# get_modified_history(15463)

def get_version(table):
    trs = table.find_all('tr')
    if len(trs)>7:
        select = trs[7].find('td').find('select')
        if select is not None:
            selectText = select.find('option',selected=True)
            if selectText is not None:
                return selectText.text.strip() # 过滤空格和换行
            else:
                return select.find_all('option')[0].text.strip()
        else:
            return trs[7].find('td').text
    else:
        return "ERROR"
        

def get_bugInfo(table, bugId):
    global reportedId  # 声明 reportedId 为全局变量
    global reported_data  # 声明 reported_data 为全局变量

    td1 = table.find('td', attrs={'id': 'bz_show_bug_column_1'})
    td2 = table.find('td', attrs={'id': 'bz_show_bug_column_2'})
    # 分析第一部分
    sub_table1 = td1.find('table')
    version = get_version(sub_table1)
    platform = sub_table1.find_all('tr')[8].find('td').find('select', attrs={'id':'rep_platform'}).find('option',selected=True).text.strip()
    op_sys = sub_table1.find_all('tr')[8].find('td').find('select', attrs={'id':'op_sys'}).find('option',selected=True).text.strip()
    priority = sub_table1.find_all('tr')[10].find('td').find('select', attrs={'id':'priority'}).find('option',selected=True).text.strip()
    severity = sub_table1.find_all('tr')[10].find('td').find('select', attrs={'id':'bug_severity'}).find('option',selected=True).text.strip()
    QA = sub_table1.find_all('tr')[13].find('td').find('a',class_='email')['href'].strip().replace('mailto:','')
    # 分析第二部分
    sub_table2 = td2.find('table')
    ccList = [option['value'] for option in sub_table2.find('div', attrs={'id': 'cc_edit_area'}).find_all('option')] # 列表推导式
    reported = [reportedId,
                sub_table2.find_all('tr')[0].find('a',class_='email')['href'].strip().replace('mailto:',''),
                sub_table2.find_all('tr')[0].find('td').text.strip().split('by')[0].strip(),
                bugId]
    reported_data.append(reported)
    reportedId+=1
    modified = get_modified_history(bugId)
    if modified is not None:
        modified_data.extend(modified)
    return [version,platform, op_sys,priority, severity, QA, ccList, reported[0]]


def get_comment_table(soup):
    table = soup.find('table', attrs={'class': 'bz_comment_table'})
    return table

def get_comments(table, bugId):
    global commentId # 声明 commentId 为全局变量

    comment_divs = table.find_all('div',class_='bz_comment')
    comments = []
    for comment_div in comment_divs:
        # 评论者 时间
        commentator = comment_div.find_all('div')[0].find('span',class_= 'bz_comment_user').find('a',class_='email')['href'].strip().replace('mailto:','')
        time = comment_div.find_all('div')[0].find('span',class_= 'bz_comment_time').text.strip()
        # 评论内容
        content = comment_div.find('pre', class_='bz_comment_text').text
        comments.append([commentId,commentator,time,content,bugId])
        commentId+=1
    return comments

def is_invalid_bug_id_page(soup):
    error_message = soup.find('title')
    if error_message and error_message.text.strip() == "Missing Bug ID":
        return True
    return False

def get_bug_details(bugId):
    url = f"https://bugzilla.yoctoproject.org/show_bug.cgi?id={bugId}"
    soup = get_html(url)
    if checkLogin('zcj15984304723@163.com',soup) is False:
        print("未正常登录",url)
        return [],[]

    if soup:
        if is_invalid_bug_id_page(soup):
            print("无效的 bug ID")
            return None, None

        bugInfo_table = get_bugInfo_table(soup)
        bugInfo = []
        if bugInfo_table:
            bugInfo = get_bugInfo(bugInfo_table, bugId)
        else:
            print("未找到Info表格")

        comments_table = get_comment_table(soup)
        comments = []
        if comments_table:
            comments = get_comments(comments_table, bugId)
        else:
            print("未找到评论表格")

        return bugInfo, comments
    else:
        print("获取页面失败")
        return None, None

# def get_bug_data(url):
#     global comment_data
#     soup = get_html(targetUrl)
#     if checkLogin('zcj15984304723@163.com', soup) is False:
#         print("未正常登录", url)
#         return [], []
#     if soup is not None:
#         table = get_table(soup)
#         if table:
#             trs = get_trs(table)

#             # 打开现有的 Excel 文件 获取现有工作表
#             bug_workbook = None
#             bug_sheet = None
#             bug_comment_workbook = None
#             bug_comment_sheet = None
#             if os.path.exists("bug_data.xlsx"):
#                 try:
#                     bug_workbook = openpyxl.load_workbook("bug_data.xlsx")
#                     bug_sheet = bug_workbook["Sheet1"]
#                 except:
#                     print("bug_data.xlsx 文件损坏，将重新创建新文件")
#                     os.remove("bug_data.xlsx")
#                     bug_workbook = openpyxl.Workbook()
#                     bug_sheet = bug_workbook.active
#                     bug_sheet.title = "Sheet1" # type: ignore
#             else:
#                 bug_workbook = openpyxl.Workbook()
#                 bug_sheet = bug_workbook.active
#                 bug_sheet.title = "Sheet1" # type: ignore

#             if os.path.exists("bug_comment_data.xlsx"):
#                 try:
#                     bug_comment_workbook = openpyxl.load_workbook("bug_comment_data.xlsx")
#                     bug_comment_sheet = bug_comment_workbook["Sheet1"]
#                 except:
#                     print("bug_comment_data.xlsx 文件损坏，将重新创建新文件")
#                     os.remove("bug_comment_data.xlsx")
#                     bug_comment_workbook = openpyxl.Workbook()
#                     bug_comment_sheet = bug_comment_workbook.active
#                     bug_comment_sheet.title = "Sheet1" # type: ignore
#             else:
#                 bug_comment_workbook = openpyxl.Workbook()
#                 bug_comment_sheet = bug_comment_workbook.active
#                 bug_comment_sheet.title = "Sheet1" # type: ignore
#             print(bug_sheet.max_row,bug_comment_sheet.max_row) # type: ignore

#             for tr in trs:
#                 bug_data = []
#                 comment_data = []
#                 tmp_data = analyse_tr(tr)
#                 bug_id = tmp_data[0]
#                 bugInfo, comments = get_bug_details(int(bug_id))
#                 if bugInfo is not None:
#                     tmp_data += bugInfo  # type: ignore # 将 bug 信息添加到临时数据列表
#                     if len(tmp_data) == 14:
#                         bug_data.append(tmp_data)
#                         # 将新的数据添加到 DataFrame 中
#                         df_bug_new = pd.DataFrame(bug_data, columns=["ID", "Product", "Component", "Assignee", "Status", "Summary", "Version", "Platform", "Op_sys", "Priority", "Severity", "QA", "CCList", "ReportedId"])
#                         # 将新的数据追加到现有的 Excel 表格中
#                         with pd.ExcelWriter("bug_data.xlsx", mode="a", engine="openpyxl", if_sheet_exists='overlay') as writer:
#                             df_bug_new.to_excel(writer, index=False, header=False, sheet_name="Sheet1", startrow=(bug_sheet.max_row if bug_sheet is not None else 0)) # type: ignore

#                 if comments is not None:
#                     if len(comments) == 5:
#                         comment_data.extend(comments)
#                         df_comment_new = pd.DataFrame(comment_data, columns=["ID", "Commentator", "Time", "Content", "BugId"])
#                         with pd.ExcelWriter("bug_comment_data.xlsx", mode="a", engine="openpyxl",if_sheet_exists='overlay') as writer:
#                             df_comment_new.to_excel(writer, index=False, header=False, sheet_name="Sheet1", startrow=(bug_comment_sheet.max_row if bug_comment_sheet is not None else 0)) # type: ignore

#                 bug_workbook.save("bug_data.xlsx")
#                 bug_comment_workbook.save("bug_comment_data.xlsx")
#                 print("Bug:", bug_id, "Finished")
#             print("数据存储成功")
#             # 保存并关闭 Excel 文件
#             bug_workbook.save("bug_data.xlsx")
#             bug_workbook.close()
#             bug_comment_workbook.save("bug_comment_data.xlsx")
#             bug_comment_workbook.close()
#         else:
#             print("未找到表格")
#     else:
#         print("获取页面失败")

#             # df_bug = pd.DataFrame(bug_data, columns=["ID", "Product", "Component", "Assignee", "Status", "Summary", "Version","Platform","Op_sys","Priority","Severity","QA","CCList", "ReportedId"])
#             # df_bug.to_excel("bug_data.xlsx", index=False)
#             # df_comment = pd.DataFrame(comment_data, columns=["ID", "Commentator", "Time", "Content", "BugId"])
#             # df_comment.to_excel("bug__comment.xlsx", index=False)
#             # df_reported = pd.DataFrame(reported_data, columns=["ID", "User", "Time", "BugId"])
#             # df_reported.to_excel("bug__reported.xlsx", index=False)
#             # df_modified = pd.DataFrame(modified_data, columns=["ID", "User", "Time", "BugId"])
#             # df_modified.to_excel("bug__modified.xlsx", index=False)

def get_bug_data(url):
    global comment_data
    soup = get_html(targetUrl)
    if checkLogin('zcj15984304723@163.com', soup) is False:
        print("未正常登录", url)
        return [], []
    if soup is not None:
        table = get_table(soup)
        if table:
            trs = get_trs(table)
            bug_data = []
            comment_data = []
            for tr in trs:
                tmp_data = analyse_tr(tr)
                bug_id = tmp_data[0]
                bugInfo, comments = get_bug_details(int(bug_id))
                if bugInfo is not None:
                    tmp_data += bugInfo  # type: ignore # 将 bug 信息添加到临时数据列表
                    if len(tmp_data) == 14:
                        bug_data.append(tmp_data)
                if comments is not None:
                    if len(comments) == 5:
                        comment_data.extend(comments)
                print("Bug:", bug_id, "Finished")
            df_bug = pd.DataFrame(bug_data, columns=["ID", "Product", "Component", "Assignee", "Status", "Summary", "Version","Platform","Op_sys","Priority","Severity","QA","CCList", "ReportedId"])
            df_bug.to_excel("bug_data.xlsx", index=False)
            df_comment = pd.DataFrame(comment_data, columns=["ID", "Commentator", "Time", "Content", "BugId"])
            df_comment.to_excel("bug__comment.xlsx", index=False)
            df_reported = pd.DataFrame(reported_data, columns=["ID", "User", "Time", "BugId"])
            df_reported.to_excel("bug__reported.xlsx", index=False)
            df_modified = pd.DataFrame(modified_data, columns=["ID", "User", "Time", "BugId"])
            df_modified.to_excel("bug__modified.xlsx", index=False)
            print("数据存储成功")
        else:
            print("未找到表格")
    else:
        print("获取页面失败")



def checkLogin(user,soup):
    if soup is not None:
        header = soup.find('div',attrs={'id':'header'})
        if header is not None:
            links = header.find('div',attrs={'id':'common_links'}).find('ul',class_='links').find_all('li')
            targetId = len(links) -1
            name = links[targetId].text.strip().split()[3].strip()
            return user == name
        else:
            print("没有找到header")
            return False
    else:
        print("获取soup失败")
        return False

get_bug_data(targetUrl)
# get_bug_details(13364)

# df_reported = pd.DataFrame(reported_data, columns=["ID", "User", "Time", "BugId"])
# df_reported.to_excel("bug__reported.xlsx", index=False)
# df_modified = pd.DataFrame(modified_data, columns=["ID", "User", "Time", "BugId"])
# df_modified.to_excel("bug__modified.xlsx", index=False)